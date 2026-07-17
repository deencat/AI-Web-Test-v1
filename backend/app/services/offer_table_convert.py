"""Convert DT offer spreadsheets to ReqIQ-compatible markdown (ReqIQ rejects .xlsx)."""
from __future__ import annotations

import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

_OFFER_SPREADSHEET_EXT = {".xlsx", ".xls", ".csv"}
_MAX_SHEET_ROWS = 120
_MAX_CELL_CHARS = 200


def is_offer_spreadsheet(filename: str) -> bool:
    return Path(filename or "").suffix.lower() in _OFFER_SPREADSHEET_EXT


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().replace("\n", " ").replace("|", "\\|")
    if len(text) > _MAX_CELL_CHARS:
        return text[: _MAX_CELL_CHARS - 1] + "…"
    return text


def _rows_to_markdown_table(rows: list[list[str]]) -> list[str]:
    if not rows:
        return ["_Empty sheet._"]
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    header = norm[0]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in norm[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return lines


def convert_offer_spreadsheet_to_markdown(content: bytes, filename: str) -> str:
    """Turn xlsx/xls/csv into markdown tables for ReqIQ upload + wiki compile."""
    ext = Path(filename or "").suffix.lower()
    stem = Path(filename or "offer").stem
    lines = [
        f"# DT Offer table — {Path(filename).name}",
        "",
        "_Auto-converted from spreadsheet for wiki compile (plan codes, promo dates, 5GBB rows)._",
        "",
    ]

    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [[_cell_str(c) for c in row] for row in reader]
        rows = [r for r in rows if any(c.strip() for c in r)]
        if len(rows) > _MAX_SHEET_ROWS:
            rows = rows[:_MAX_SHEET_ROWS]
            lines.append(f"_Truncated to {_MAX_SHEET_ROWS} rows._")
            lines.append("")
        lines.append("## Data")
        lines.append("")
        lines.extend(_rows_to_markdown_table(rows))
        return "\n".join(lines).strip() + "\n"

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel offer tables") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}")
            lines.append("")
            rows: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= _MAX_SHEET_ROWS:
                    lines.append(f"_Truncated to {_MAX_SHEET_ROWS} rows._")
                    lines.append("")
                    break
                cells = [_cell_str(v) for v in row]
                if any(cells):
                    rows.append(cells)
            lines.extend(_rows_to_markdown_table(rows))
            lines.append("")
    finally:
        wb.close()

    if len(lines) <= 4:
        raise ValueError(f"No readable data in {filename}")

    return "\n".join(lines).strip() + "\n"


def reqiq_markdown_filename(original_filename: str) -> str:
    stem = Path(original_filename or "offer").stem
    return f"{stem} (offer table).md"
